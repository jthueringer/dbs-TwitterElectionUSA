# *.xlsx laden und lesen
from openpyxl import load_workbook
# *.xlsx schreiben
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
# hashtags finden und schneiden
import re 
regex = re.compile('[^a-zA-Z_0-9]') 

# workbook mit unbereinigten Daten laden
wb = load_workbook('american-election-tweets.xlsx')
ws = wb['american-election-tweets']
wslen = ws.max_row-1

# wb fuer bereinigte Daten zum Schreiben erstellen
wb_new = Workbook()
dest_filename = 'american-election-repair.xlsx'
ws_new = wb_new.active
ws_new.title = 'bereinigt'

# Hilfsfunktion ist_enthalten fuer tag
def ist_enthalten (cell_value, liste):
	lenListe = len(liste)
	for name in range(0,lenListe):
		if cell_value==liste[name]:
			return True
		else:
			continue
		return False

# fuer Relation hashtag:
hashtag_list = []
zaehlerID = 0
ws_new['F1'] = 'text'
ws_new['G1'] = 'hashtag'
for row in range(0, wslen):
	current_hash_list = ''
	hashtag_id_list = ''
	temp=[]
	cell_text_read = "{col}{row}".format(col='B', row=(row+2))
	cell_text_write = "{col}{row}".format(col='F', row=(row+2))
	cell_hash_write = "{col}{row}".format(col='G', row=(row+2))
	tags = str(ws[cell_text_read].value)
	cell_read = str(ws[cell_text_read].value)
	
	# text in wb_new schreiben
	ws_new[cell_text_write] = cell_read
	
	# zelleninhalt teilen und #finden, wie auch entfernen
	temp=[tag.strip("#") for tag in tags.split() if tag.startswith("#")]
	if (len(temp)>0):
		temp0 = temp[0].lower()
		temp0 = regex.sub('', temp0)
		if (temp0 == ''):
			pass
		elif (zaehlerID==0):
			temp0 = regex.sub('', temp0)
			hashtag_list = [temp0]
			current_hash_list += str(temp0)
			zaehlerID+=1
		elif (zaehlerID>0):
			if ist_enthalten(temp0, hashtag_list)==True:
				for id in range(0,len(hashtag_list)):
					if (hashtag_list[id] == temp0):
						current_hash_list += str(temp0)
			else:
				hashtag_list.append(temp0)
				current_hash_list += str(temp0)
	if (len(temp)>1):
		for i in range(1,len(temp)):
			tempi = temp[i].lower()
			tempi = regex.sub('', tempi)
			if (tempi == ''):
				pass
			elif ist_enthalten(tempi, hashtag_list)==True:
				for id in range(0,len(hashtag_list)):
					if (hashtag_list[id] == tempi):
						current_hash_list += ' ' + str(tempi)
			else:
				hashtag_list.append(tempi)
				current_hash_list += ' ' + str(tempi)
	if (hashtag_list == ''):
		ws_new[cell_hash_write] = ''
	else:
		ws_new[cell_hash_write] = current_hash_list

# nur Auflistung der Namen
ws_new['H1'] = 'tag'		
for tag in range(0,len(hashtag_list)):
	cell_tag_write = "{col}{row}".format(col='H', row=(tag+2))
	ws_new[cell_tag_write] = str(hashtag_list[tag])
	
	
# fuer Relation tweet:
tweet_list = []
zaehlerID = 0
ws_new['A1'] = 'handle'
ws_new['B1'] = 'timeTweeted'
ws_new['C1'] = 'favourite_count'
ws_new['D1'] = 'retweet_count'
ws_new['E1'] = 'original_author'
for row in range(0, wslen):
	current_hash_list = ''
	hashtag_id_list = ''
	temp=[]
	cell_handle_read = "{col}{row}".format(col='A', row=(row+2))
	cell_handle_write = "{col}{row}".format(col='A', row=(row+2))
	ws_new[cell_handle_write] = ws[cell_handle_read].value
	
	cell_time_read = "{col}{row}".format(col='E', row=(row+2))
	cell_time_write = "{col}{row}".format(col='B', row=(row+2))
	time = ws[cell_time_read].value.replace('T', ' ')
	ws_new[cell_time_write] = time		
	
	cell_fav_read = "{col}{row}".format(col='I', row=(row+2))
	cell_fav_write = "{col}{row}".format(col='C', row=(row+2))
	ws_new[cell_fav_write] = ws[cell_fav_read].value
	
	cell_retw_read = "{col}{row}".format(col='H', row=(row+2))
	cell_retw_write = "{col}{row}".format(col='D', row=(row+2))
	ws_new[cell_retw_write] = ws[cell_retw_read].value
	
	cell_auth_read = "{col}{row}".format(col='D', row=(row+2))
	cell_auth_write = "{col}{row}".format(col='E', row=(row+2))
	kl=ws[cell_auth_read].value
	if (str(kl) == 'None'):
		auth = ''
	else:
		auth = str(ws[cell_auth_read].value)
	ws_new[cell_auth_write] = auth
	
	
wb_new.save(filename = dest_filename)
