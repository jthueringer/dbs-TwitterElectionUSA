# *.xlsx laden und lesen
from openpyxl import load_workbook
# Verbindung zur Datenbank und string-escape fuer sql
from pgdb import connect
try:
	pg=connect(dbname='election', user='julez', password='haha')
except:
	print('keine db-verbindung')
# workbook mit aufbereiteten Daten laden
wb = load_workbook('american-election-repair.xlsx')
ws = wb['bereinigt']
wslen = ws.max_row-1


# insert in relation tweet
cursor = pg.cursor()
for row in range(0, wslen):
	cell_handle = "{col}{row}".format(col='A', row=(row+2))
	cell_time = "{col}{row}".format(col='B', row=(row+2))
	cell_fav = "{col}{row}".format(col='C', row=(row+2))
	cell_retw = "{col}{row}".format(col='D', row=(row+2))
	cell_text = "{col}{row}".format(col='F', row=(row+2))
	cell_auth = "{col}{row}".format(col='E', row=(row+2))
	
	format_str = "INSERT INTO tweet VALUES (%s, %s, %s, %s, %s, %s)"
	data_tweet = [ws[cell_handle].value, ws[cell_time].value, ws[cell_fav].value, ws[cell_retw].value, ws[cell_text].value, ws[cell_auth].value]
	cursor.execute(format_str, data_tweet)
pg.commit()
cursor.close()


# insert in relation hashtag
cursor = pg.cursor()
for row in range(0, wslen):
	cell_tag = "{col}{row}".format(col='H', row=(row+2))
	if (str(ws[cell_tag].value) == 'None'):
		break
	else:
		data_tag = [ws[cell_tag].value]
	
	format_str = "INSERT INTO hashtag VALUES (%s)"
	cursor.execute(format_str, data_tag)
pg.commit()
cursor.close()

# insert in relation has
cursor = pg.cursor()
for row in range(0, wslen):
	cell_time = "{col}{row}".format(col='B', row=(row+2))
	cell_handle = "{col}{row}".format(col='A', row=(row+2))
	cell_tag = "{col}{row}".format(col='G', row=(row+2))
	tag_list = str(ws[cell_tag].value)
	
	if (tag_list == 'None'):
		continue
	else:
		for tag in tag_list.split():
			format_str = "INSERT INTO has VALUES (%s, %s, %s)"
			data_has = [ws[cell_time].value, ws[cell_handle].value, tag]
			cursor.execute(format_str, data_has)
pg.commit()
cursor.close()
